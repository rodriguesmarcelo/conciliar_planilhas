# -*- coding: utf-8 -*-
"""Helpers dos testes da Task 09 (paridade / bordas / interop).

Gera fixtures .xlsx por código, roda o caminho do desktop (core/) e o caminho
web (router real) e compara os workbooks resultantes.
"""
import os
import shutil
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import Workbook, load_workbook
from starlette.datastructures import Headers, UploadFile

from core.engine import run as engine_run
from core.exporter import export_result
from core.reader import read_sheet, validate_columns
from web.modules.conciliador import storage
from web.modules.conciliador.router import run_reconciliation


def make_xlsx(path, start_row, data_rows, n_cols):
    """Cria um .xlsx com (start_row-1) linhas de cabeçalho 'lixo' (abrangendo
    n_cols colunas) e os dados a partir de start_row.

    `data_rows`: lista de linhas; cada linha é a lista dos valores das colunas
    1..N (1-based no layout do perfil). Valores None viram célula vazia.
    """
    wb = Workbook()
    ws = wb.active
    # Cabeçalho lixo: garante que max_column == n_cols mesmo sem dados.
    for r in range(1, start_row):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c, value=f"h{r}.{c}")
    for i, row in enumerate(data_rows):
        excel_row = start_row + i
        for c, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=excel_row, column=c, value=val)
    wb.save(path)
    return path


def run_desktop_path(profile, path_a, path_b, out_path):
    """Espelha ui/screen_run._run_process: validate → read → run → export."""
    sheets = profile["sheets"]
    mode = profile.get("mode", "dual")
    date_fmt = profile.get("date_format", "dd/mm/yyyy")

    validate_columns(path_a, sheets[0])
    if mode == "dual" and len(sheets) > 1:
        validate_columns(path_b, sheets[1])

    df_a = read_sheet(path_a, sheets[0], date_fmt)
    df_b = read_sheet(path_b, sheets[1], date_fmt) if (mode == "dual" and path_b) else None
    result = engine_run(profile, df_a, df_b)
    export_result(result, profile, out_path, file_a=path_a, file_b=path_b)
    return out_path, result


def make_upload(path, filename=None):
    """Constrói um starlette UploadFile a partir de um arquivo em disco."""
    if path is None:
        return None
    return UploadFile(
        file=open(path, "rb"),
        filename=filename or os.path.basename(path),
        headers=Headers({}),
    )


def run_web_path(profile_id, path_a, path_b, copy_to):
    """Chama o router web real e copia o .xlsx de resultado para copy_to."""
    ua = make_upload(path_a)
    ub = make_upload(path_b)
    try:
        resp = run_reconciliation(profile_id, ua, ub)
    finally:
        if ua is not None:
            ua.file.close()
        if ub is not None:
            ub.file.close()
    src = storage.result_path(resp.download_token)
    assert src, "resultado web não encontrado no storage"
    shutil.copyfile(src, copy_to)
    return copy_to, resp


def compare_workbooks(path_x, path_y, ignore_sheets=("Resumo",)):
    """Compara abas de dados célula a célula (ignora a aba volátil 'Resumo')."""
    wx = load_workbook(path_x, data_only=True)
    wy = load_workbook(path_y, data_only=True)
    sx = [s for s in wx.sheetnames if s not in ignore_sheets]
    sy = [s for s in wy.sheetnames if s not in ignore_sheets]
    assert sx == sy, f"Abas diferentes: {sx} != {sy}"
    for name in sx:
        rows_x = list(wx[name].iter_rows(values_only=True))
        rows_y = list(wy[name].iter_rows(values_only=True))
        assert rows_x == rows_y, (
            f"Aba '{name}' difere entre desktop e web:\n{rows_x}\n!=\n{rows_y}"
        )


def counts_of(result):
    """Contagens normalizadas a partir do dict de resultado do engine."""
    def n(df):
        return 0 if df is None else len(df)
    if result.get("normalizados") is not None:
        return {"normalizados": n(result.get("normalizados"))}
    return {
        "conciliados": n(result.get("conciliados")),
        "nao_conciliados_a": n(result.get("nao_conciliados_a")),
        "nao_conciliados_b": n(result.get("nao_conciliados_b")),
    }


def run_suite(tests):
    """Executa uma lista de (nome, fn) imprimindo PASS/FAIL. Sai !=0 se falhar."""
    failed = 0
    for name, fn in tests:
        try:
            fn()
            print(f"  PASS  {name}")
        except Exception as e:
            failed += 1
            print(f"  FAIL  {name}: {type(e).__name__}: {e}")
    total = len(tests)
    print(f"\n{total - failed}/{total} testes passaram.")
    sys.exit(1 if failed else 0)
