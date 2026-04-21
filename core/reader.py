# -*- coding: utf-8 -*-
import re
import os
from datetime import datetime, date
from typing import Optional

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Exceções
# ---------------------------------------------------------------------------

class ColumnNotFoundError(Exception):
    def __init__(self, col_index: int, sheet_label: str):
        self.col_index = col_index
        self.sheet_label = sheet_label
        super().__init__(
            f"A coluna {col_index} não foi encontrada na planilha '{sheet_label}'. "
            f"Verifique o perfil ou o arquivo selecionado."
        )


class InvalidFileError(Exception):
    pass


# ---------------------------------------------------------------------------
# Transformações individuais
# ---------------------------------------------------------------------------

def _transform_remove_newlines(value: str) -> str:
    return value.replace('\n', ' ').replace('\r', '')


def _transform_remove_semicolons(value: str) -> str:
    return value.replace(';', '')


def _transform_remove_dots(value: str) -> str:
    return value.replace('.', '')


def _transform_remove_commas(value: str) -> str:
    """Converte vírgula decimal BR para ponto."""
    return value.replace(',', '.')


def _transform_remove_special_chars(value: str) -> str:
    return re.sub(r'\D', '', value)


def _transform_to_int(value: str) -> str:
    s = str(value).strip()
    # Remover espaços não-quebráveis
    s = re.sub(r'[\s\xa0]', '', s)
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s


def _transform_remove_cd_suffix(value: str) -> str:
    """Remove letras C/D e espaços (incluindo \xa0) ao final do valor."""
    return re.sub(r'[\s\xa0]*[CDcd]$', '', str(value)).strip()


def _transform_remove_asterisk_suffix(value: str) -> str:
    """Remove * e espaços ao final (valores como '0,00 *')."""
    return re.sub(r'[\s\xa0]*\*$', '', str(value)).strip()


def _transform_absolute_value(value: str) -> str:
    try:
        clean = re.sub(r'[\s\xa0]', '', str(value))
        return str(abs(float(clean)))
    except (ValueError, TypeError):
        return value


def _transform_strip_after_dash(value: str) -> str:
    return value.split('-')[0].strip()


def _transform_strip_after_slash(value: str) -> str:
    return value.split('/')[0].strip()


def _transform_lstrip_zeros(value: str) -> str:
    stripped = value.lstrip('0')
    return stripped if stripped else '0'


def _transform_apply_cnpj_mask(value: str) -> str:
    digits = re.sub(r'\D', '', value)
    digits = digits.zfill(14)[:14]
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"


TRANSFORMATIONS = {
    'remove_newlines':     _transform_remove_newlines,
    'remove_semicolons':   _transform_remove_semicolons,
    'remove_dots':         _transform_remove_dots,
    'remove_commas':       _transform_remove_commas,
    'remove_special_chars':_transform_remove_special_chars,
    'to_int':              _transform_to_int,
    'remove_cd_suffix':    _transform_remove_cd_suffix,
    'absolute_value':      _transform_absolute_value,
    'strip_after_dash':    _transform_strip_after_dash,
    'strip_after_slash':   _transform_strip_after_slash,
    'lstrip_zeros':        _transform_lstrip_zeros,
    'apply_cnpj_mask':     _transform_apply_cnpj_mask,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False


def _parse_date(value, date_format: str = "dd/mm/yyyy") -> Optional[date]:
    """Tenta converter o valor para date. Retorna None se inválido."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # Mapear formato dd/mm/yyyy → strptime
    fmt_map = {
        'dd/mm/yyyy': '%d/%m/%Y',
        'mm/dd/yyyy': '%m/%d/%Y',
        'yyyy-mm-dd': '%Y-%m-%d',
    }
    fmt = fmt_map.get(date_format, '%d/%m/%Y')
    try:
        return datetime.strptime(s, fmt).date()
    except ValueError:
        return None


def _apply_transformations(value, transformations: list) -> str:
    """Aplica lista de transformações em sequência sobre o valor (como string)."""
    s = str(value) if value is not None else ''
    for t in transformations:
        fn = TRANSFORMATIONS.get(t)
        if fn:
            s = fn(s)
    return s


def _to_numeric(value) -> Optional[float]:
    """Tenta converter um valor para float, retornando None se falhar."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Validação de colunas
# ---------------------------------------------------------------------------

def validate_columns(filepath: str, sheet_config: dict):
    """
    Verifica se todos os col_index configurados existem na planilha.
    Lança ColumnNotFoundError se alguma coluna estiver fora do range.
    """
    label = sheet_config.get('label', 'Planilha')
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        max_col = ws.max_column or 0
        wb.close()
    except Exception as e:
        raise InvalidFileError(f"Não foi possível abrir o arquivo: {e}")

    for col_cfg in sheet_config.get('columns', []):
        col_index = col_cfg.get('col_index', 0)
        if col_index < 1 or col_index > max_col:
            raise ColumnNotFoundError(col_index, label)


# ---------------------------------------------------------------------------
# Pré-visualização
# ---------------------------------------------------------------------------

def preview_sheet(filepath: str, max_rows: int = 20) -> list:
    """
    Retorna as primeiras max_rows linhas brutas da planilha.
    Primeira linha do retorno é o cabeçalho de colunas: ['Col', 1, 2, 3, ...]
    Cada linha seguinte: [numero_linha, val1, val2, ...]
    """
    try:
        ext = os.path.splitext(filepath)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(filepath, header=None, nrows=max_rows)
            rows = df.values.tolist()
            max_col = max(len(r) for r in rows) if rows else 0
            header = ['Col'] + list(range(1, max_col + 1))
            result = [header]
            for i, row in enumerate(rows, 1):
                result.append([i] + list(row))
            return result
        else:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
            max_col = ws.max_column or 0
            header = ['Col'] + list(range(1, max_col + 1))
            result = [header]
            count = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                result.append([row_idx] + list(row))
                count += 1
                if count >= max_rows:
                    break
            wb.close()
            return result
    except Exception as e:
        raise InvalidFileError(f"Erro ao ler arquivo para pré-visualização: {e}")


# ---------------------------------------------------------------------------
# Leitura principal
# ---------------------------------------------------------------------------

def read_sheet(filepath: str, sheet_config: dict, date_format: str = "dd/mm/yyyy") -> pd.DataFrame:
    """
    Lê uma planilha conforme as configurações do perfil.
    Retorna DataFrame com colunas padronizadas + coluna '__linha__' com número real da linha.
    """
    label = sheet_config.get('label', 'Planilha')
    start_row = sheet_config.get('start_row', 1)
    columns_cfg = sheet_config.get('columns', [])

    # Carregar workbook
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        raw_df = pd.read_csv(filepath, header=None, dtype=str)
        rows_raw = []
        for idx, row in raw_df.iterrows():
            rows_raw.append((idx + 1, list(row)))
    else:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            raise InvalidFileError(f"Não foi possível abrir o arquivo '{label}': {e}")
        ws = wb.active
        rows_raw = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
            rows_raw.append((row_idx, list(row)))

    # Separar campos principais e fallbacks
    # primary_field → config do campo principal (valor_debito → campo 'valor')
    # fallback_field → config do campo fallback (valor_credito, is_fallback_of='valor')
    primary_fields = {}   # field_name -> col_cfg
    fallback_fields = {}  # parent_field_name -> col_cfg

    for col_cfg in columns_cfg:
        field = col_cfg['field']
        parent = col_cfg.get('is_fallback_of')
        if parent:
            fallback_fields[parent] = col_cfg
        else:
            primary_fields[field] = col_cfg

    # Identificar blank_limit (Modelo 4)
    blank_limit_field = None
    blank_limit_count = None
    for col_cfg in columns_cfg:
        if 'blank_limit' in col_cfg:
            blank_limit_field = col_cfg['field']
            blank_limit_count = col_cfg['blank_limit']
            break

    records = []
    consecutive_blanks = 0

    for row_idx, row_values in rows_raw:
        # Garantir tamanho mínimo da linha
        max_needed = max((c['col_index'] for c in columns_cfg), default=1)
        while len(row_values) < max_needed:
            row_values.append(None)

        record = {'__linha__': row_idx}
        skip = False

        # ── Processar campos principais (não-fallback) ──
        for field, col_cfg in primary_fields.items():
            col_idx = col_cfg['col_index'] - 1  # 0-indexed
            raw_val = row_values[col_idx] if col_idx < len(row_values) else None

            # skip_if_empty para campos de controle (ex: numero_cheque, numero_nota)
            if col_cfg.get('skip_if_empty') and _is_empty(raw_val):
                skip = True
                break

            # skip_if_invalid_date
            if col_cfg.get('skip_if_invalid_date'):
                parsed = _parse_date(raw_val, date_format)
                if parsed is None:
                    skip = True
                    break
                record[field] = parsed
                continue

            # Verificar se é campo de data (sem skip_if_invalid_date mas com nome 'data')
            # Aplicar transformações normais
            if field == 'data' and not col_cfg.get('skip_if_invalid_date'):
                parsed = _parse_date(raw_val, date_format)
                record[field] = parsed
                continue

            # Campos numéricos: tentar converter antes de transformações textuais
            transformations = col_cfg.get('transformations', [])
            multiply = col_cfg.get('multiply_minus_one', False)

            # Se não há transformações textuais, tratar como numérico direto
            if not transformations:
                num = _to_numeric(raw_val)
                if num is not None:
                    if multiply:
                        num = num * -1
                    record[field] = num
                else:
                    # Campo textual sem transformações (ex: numero_nota, cnpj, fornecedor)
                    record[field] = str(raw_val).strip() if raw_val is not None else None
            else:
                # Aplicar transformações (modo texto)
                val_str = _apply_transformations(raw_val, transformations)
                # Tentar converter para numérico após transformações
                # Se 'to_int' está nas transformações, manter como int
                if 'to_int' in transformations:
                    try:
                        num = int(val_str)
                        if multiply:
                            num = num * -1
                        record[field] = num
                    except (ValueError, TypeError):
                        record[field] = val_str if val_str else None
                else:
                    # Transformações que devem manter o resultado como string
                    _string_transforms = {'lstrip_zeros', 'strip_after_dash', 'strip_after_slash',
                                          'apply_cnpj_mask', 'remove_newlines', 'remove_semicolons',
                                          'remove_special_chars'}
                    force_string = bool(_string_transforms.intersection(set(transformations)))

                    num = _to_numeric(val_str)
                    if num is not None and not force_string:
                        if multiply:
                            num = num * -1
                        record[field] = num
                    else:
                        record[field] = val_str if val_str else None

        if skip:
            # Checar blank_limit
            if blank_limit_field is not None:
                # Só conta blank para o campo de controle do blank_limit
                bl_cfg = next((c for c in columns_cfg if c['field'] == blank_limit_field), None)
                if bl_cfg:
                    bl_val = row_values[bl_cfg['col_index'] - 1] if bl_cfg['col_index'] - 1 < len(row_values) else None
                    if _is_empty(bl_val):
                        consecutive_blanks += 1
                        if consecutive_blanks >= blank_limit_count:
                            break
                    else:
                        consecutive_blanks = 0
            continue

        # ── Processar fallbacks ──
        # Ex: valor_debito (primary) + valor_credito (fallback de 'valor')
        # Resultado: campo 'valor' = debito se preenchido, senão credito
        for parent_field, fb_cfg in fallback_fields.items():
            # Encontrar o campo primary correspondente (ex: valor_debito → is_fallback_of=valor)
            # O campo primário que gera 'valor' é valor_debito
            primary_name = None
            primary_value = None
            for f, cfg in primary_fields.items():
                if cfg.get('is_fallback_of') == parent_field:
                    # Esse campo primário já foi processado mas com nome errado
                    # Na realidade, valor_debito não tem is_fallback_of, é o valor_credito que tem
                    pass

            # O campo primary que "alimenta" parent_field: buscar campo sem is_fallback_of
            # cujo resultado já está em record com nome de campo like 'valor_debito'
            # Lógica: parent_field='valor' → buscar 'valor_debito' em record
            debito_key = f"{parent_field}_debito"
            credito_key = f"{parent_field}_credito"

            debito_val = record.get(debito_key)
            fb_col_idx = fb_cfg['col_index'] - 1
            fb_raw = row_values[fb_col_idx] if fb_col_idx < len(row_values) else None

            transformations = fb_cfg.get('transformations', [])
            multiply = fb_cfg.get('multiply_minus_one', False)

            if not transformations:
                fb_num = _to_numeric(fb_raw)
                if multiply and fb_num is not None:
                    fb_num = fb_num * -1
            else:
                fb_str = _apply_transformations(fb_raw, transformations)
                fb_num = _to_numeric(fb_str)
                if fb_num is not None and multiply:
                    fb_num = fb_num * -1

            # Regra de fallback: se debito vazio/zero, usar credito
            if _is_empty(debito_val) or debito_val == 0.0:
                record[parent_field] = fb_num
            else:
                record[parent_field] = debito_val

            # Remover chaves auxiliares (valor_debito, valor_credito)
            record.pop(debito_key, None)
            record.pop(credito_key, None)

        # Resetar contador de brancos consecutivos
        consecutive_blanks = 0
        records.append(record)

    df = pd.DataFrame(records)

    # Converter colunas com 'to_int' nas transformações para Int64 nullable
    for col_cfg in columns_cfg:
        if 'to_int' in col_cfg.get('transformations', []):
            field = col_cfg['field']
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce').astype(pd.Int64Dtype())

    if df.empty:
        # Garantir colunas mínimas
        cols = ['__linha__'] + [c['field'] for c in columns_cfg if not c.get('is_fallback_of')]
        # Adicionar campos de fallback como parent_field
        for pf in fallback_fields:
            if pf not in cols:
                cols.append(pf)
        df = pd.DataFrame(columns=cols)

    return df
