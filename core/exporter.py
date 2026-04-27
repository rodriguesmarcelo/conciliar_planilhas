# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from datetime import datetime, date
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constantes de formatação
# ---------------------------------------------------------------------------

HEADER_FONT       = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL       = PatternFill("solid", fgColor="1F497D")   # Azul escuro
HEADER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT         = Font(name="Arial", size=10)
ALT_FILL          = PatternFill("solid", fgColor="F2F2F2")   # Cinza claro
WHITE_FILL        = PatternFill("solid", fgColor="FFFFFF")
CENTER_ALIGN      = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN        = Alignment(horizontal="left",   vertical="center")

DATE_FORMAT       = "DD/MM/YYYY"
NUMBER_FORMAT     = '#,##0.00'
INTEGER_FORMAT    = '0'

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9")
)

COL_MIN_WIDTH = 10
COL_MAX_WIDTH = 60

# Palavras-chave que identificam colunas de data, valor e linha pelo header
_DATE_KEYWORDS   = {"data", "data_do_pagamento", "data_pagamento", "date"}
_VALUE_KEYWORDS  = {"valor", "valor_da_parcela", "valor_parcela", "juros", "multa",
                    "desconto", "value", "vl"}
_LINE_KEYWORDS   = {"linha", "line", "row"}


def _is_date_col(header: str) -> bool:
    h = header.lower().replace(" ", "_").replace("ã", "a").replace("é", "e")
    return any(kw in h for kw in _DATE_KEYWORDS)


def _is_value_col(header: str) -> bool:
    h = header.lower().replace(" ", "_")
    return any(kw in h for kw in _VALUE_KEYWORDS)


def _is_line_col(header: str) -> bool:
    h = header.lower().replace(" ", "_").replace("ã", "a")
    return any(kw in h for kw in _LINE_KEYWORDS)


# ---------------------------------------------------------------------------
# Escrita de uma aba de dados
# ---------------------------------------------------------------------------

def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    """Escreve um DataFrame em uma aba do workbook com formatação completa."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel limita nome a 31 chars

    if df is None or df.empty:
        # Aba vazia com apenas o cabeçalho
        if df is not None and len(df.columns) > 0:
            for col_idx, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font  = HEADER_FONT
                cell.fill  = HEADER_FILL
                cell.alignment = HEADER_ALIGN
        ws.freeze_panes = "A2"
        return

    headers = list(df.columns)
    n_cols  = len(headers)

    # ── Cabeçalho ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # ── Dados ──
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        fill = WHITE_FILL if (row_idx % 2 == 0) else ALT_FILL

        for col_idx, header in enumerate(headers, 1):
            raw_val = row[header]

            # Normalizar valor
            val = _normalize_value(raw_val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = DATA_FONT
            cell.fill   = fill
            cell.border = THIN_BORDER

            # Aplicar formato conforme tipo de coluna
            if _is_date_col(header) and val is not None:
                cell.number_format = DATE_FORMAT
                cell.alignment     = CENTER_ALIGN
            elif _is_line_col(header):
                cell.number_format = INTEGER_FORMAT
                cell.alignment     = CENTER_ALIGN
            elif _is_value_col(header) and isinstance(val, (int, float)):
                cell.number_format = NUMBER_FORMAT
                cell.alignment     = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    # ── Largura automática das colunas ──
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))

        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                # Para datas, o formato dd/mm/yyyy tem 10 chars
                if isinstance(cell_val, (datetime, date)):
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(str(cell_val)))

        adjusted = max(COL_MIN_WIDTH, min(COL_MAX_WIDTH, max_len + 2))
        ws.column_dimensions[col_letter].width = adjusted


def _normalize_value(val):
    """Converte valores pandas para tipos nativos Python aceitos pelo openpyxl."""
    if val is None:
        return None
    # pd.NA ou pd.NaT
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    # date / datetime — retornar como datetime para openpyxl formatar
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime(val.year, val.month, val.day)
    # float nativo — deve vir antes da tentativa de int
    if isinstance(val, float):
        return val
    # string nativa
    if isinstance(val, str):
        return val
    # int nativo
    if isinstance(val, int):
        return val
    # numpy int / float
    try:
        import numpy as np
        if isinstance(val, np.integer):
            return int(val)
        if isinstance(val, np.floating):
            return float(val)
    except ImportError:
        pass
    # pandas Int64 nullable (não é int/float nativo)
    try:
        iv = int(val)
        # Só retorna int se a conversão for exata (sem perda decimal)
        if float(iv) == float(val):
            return iv
        return float(val)
    except (TypeError, ValueError, OverflowError):
        pass
    return val


# ---------------------------------------------------------------------------
# Aba de Resumo
# ---------------------------------------------------------------------------

def _write_resumo(wb: Workbook, profile: dict, result: dict,
                  file_a: Optional[str], file_b: Optional[str],
                  executed_at: datetime):
    """Cria a aba 'Resumo' como primeira aba do workbook."""
    ws = wb.create_sheet(title="Resumo")

    title_font  = Font(name="Arial", size=14, bold=True, color="1F497D")
    label_font  = Font(name="Arial", size=10, bold=True)
    value_font  = Font(name="Arial", size=10)
    section_fill = PatternFill("solid", fgColor="DDEEFF")

    def write_row(row, col, label, value=None, *, label_font=label_font, value_font=value_font):
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = label_font
        if value is not None:
            vc = ws.cell(row=row, column=col + 1, value=value)
            vc.font = value_font

    # Título
    ws.cell(row=1, column=1, value="Conciliador de Planilhas — Relatório de Resultados").font = title_font
    ws.merge_cells("A1:D1")

    # Informações gerais
    row = 3
    write_row(row, 1, "Perfil:", profile.get("name", "—")); row += 1
    write_row(row, 1, "Modo:", "Dual (2 planilhas)" if profile.get("mode") == "dual" else "Single (1 planilha)"); row += 1
    write_row(row, 1, "Data/Hora da Execução:", executed_at.strftime("%d/%m/%Y %H:%M:%S")); row += 1

    if file_a:
        write_row(row, 1, "Arquivo A:", os.path.basename(file_a)); row += 1
    if file_b:
        write_row(row, 1, "Arquivo B:", os.path.basename(file_b)); row += 1

    row += 1
    # Contagens
    ws.cell(row=row, column=1, value="Contagem de Registros").font = Font(name="Arial", size=11, bold=True, color="1F497D")
    row += 1

    mode = profile.get("mode", "dual")
    if mode == "single":
        norm = result.get("normalizados")
        write_row(row, 1, "Registros normalizados:", len(norm) if norm is not None else 0); row += 1
    else:
        conc = result.get("conciliados")
        nca  = result.get("nao_conciliados_a")
        ncb  = result.get("nao_conciliados_b")
        label_a = profile["sheets"][0]["label"] if profile.get("sheets") else "A"
        label_b = profile["sheets"][1]["label"] if len(profile.get("sheets", [])) > 1 else "B"
        write_row(row, 1, "Registros conciliados:",                len(conc) if conc is not None else 0); row += 1
        write_row(row, 1, f"Não conciliados — {label_a}:",         len(nca)  if nca  is not None else 0); row += 1
        write_row(row, 1, f"Não conciliados — {label_b}:",         len(ncb)  if ncb  is not None else 0); row += 1

    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def export_result(
    result: dict,
    profile: dict,
    output_path: str,
    file_a: Optional[str] = None,
    file_b: Optional[str] = None,
) -> str:
    """
    Gera arquivo Excel de resultado.

    Args:
        result:      retorno de engine.run()
        profile:     perfil completo
        output_path: caminho completo do arquivo de saída
        file_a:      caminho do arquivo A (para o resumo)
        file_b:      caminho do arquivo B (para o resumo)

    Returns:
        output_path confirmado
    """
    executed_at = datetime.now()
    mode = profile.get("mode", "dual")
    output_tabs = profile.get("output", {}).get("tabs", [])

    wb = Workbook()
    # Remover aba padrão criada pelo openpyxl
    wb.remove(wb.active)

    # ── Abas de dados primeiro ──
    if mode == "single":
        tab = next((t for t in output_tabs if t["source"] == "normalizados"), None)
        tab_name = tab["name"] if tab else "Dados Normalizados"
        _write_sheet(wb, tab_name, result.get("normalizados"))

    else:
        # Mapa source → DataFrame
        source_map = {
            "conciliados":       result.get("conciliados"),
            "nao_conciliados_a": result.get("nao_conciliados_a"),
            "nao_conciliados_b": result.get("nao_conciliados_b"),
        }

        for tab in output_tabs:
            source = tab["source"]
            tab_name = tab["name"]
            df = source_map.get(source)
            _write_sheet(wb, tab_name, df)

    # ── Aba de Resumo sempre por último ──
    _write_resumo(wb, profile, result, file_a, file_b, executed_at)

    # Salvar
    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path
